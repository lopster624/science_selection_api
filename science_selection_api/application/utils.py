import datetime
import re
from io import BytesIO

from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Prefetch, Count, Q, F, Case, When, Value, OuterRef
from django_filters import NumberFilter, BaseInFilter, CharFilter, AllValuesMultipleFilter
from django_filters.rest_framework import FilterSet
from docxtpl import DocxTemplate
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from rest_framework.exceptions import ValidationError, ParseError
from rest_framework.filters import OrderingFilter
from rest_framework.generics import get_object_or_404
from rest_framework.pagination import PageNumberPagination

from account.models import Member, Affiliation, Booking, BookingType
from utils import constants as const
from utils.calculations import get_current_draft_year, convert_float
from utils.constants import BOOKED, MEANING_COEFFICIENTS, PATH_TO_RATING_LIST, \
    PATH_TO_CANDIDATES_LIST, PATH_TO_EVALUATION_STATEMENT, TRUE_VALUES, FALSE_VALUES, MASTER_ROLE_NAME
from utils.constants import NAME_ADDITIONAL_FIELD_TEMPLATE
from .models import Application, AdditionField, AdditionFieldApp, MilitaryCommissariat, Competence, ViewedApplication, \
    Education, ApplicationNote


class PaginationApplication(PageNumberPagination):
    """Пагинация для списка анкет."""
    page_size = 50


def has_affiliation(member, affiliation):
    """
    Возвращает true, если у member есть принадлежность affiliation
    :param member: экземпляр класса Member
    :param affiliation: экземпляр класса Affiliation
    :return: True/False
    """
    if member.is_slave():
        if affiliation.direction.id in get_slave_affiliations_id(member):
            return True
        return False
    elif member.is_master():
        if affiliation.id in get_master_affiliations_id(member):
            return True
        return False
    raise PermissionDenied('Доступ с текущей ролью запрещен.')


def get_master_affiliations_id(member):
    """
    Получает список id принадлежностей мастера
    :param member: экземпляр класса Member, должен иметь роль master!
    :return: Список id принадлежностей
    """
    return Affiliation.objects.filter(member=member).values_list('id', flat=True)


def get_slave_affiliations_id(member):
    """
    Получает список id принадлежностей кандидата
    :param member: экземпляр класса Member, должен иметь роль slave!
    :return: Список id принадлежностей
    """
    return Member.objects.prefetch_related(
        'application__directions__id').only('application__directions__id').values_list(
        'application__directions__id', flat=True).distinct().filter(pk=member.pk)


def get_booked_type():
    """ Возвращает объект BookingType типа бронирования 'Отобран' """
    return BookingType.objects.get(name=const.BOOKED)


def get_in_wishlist_type():
    """ Возвращает объект BookingType типа бронирования 'В избранном' """
    return BookingType.objects.get(name=const.IN_WISHLIST)


def is_master(user):
    """
    Проверяет, является ли пользователь отбирающим.
    :param user: объект user
    :return: bool
    """
    try:
        return user.member.is_master()
    except AttributeError:
        return False


def is_slave(user):
    """
    Проверяет, является ли пользователь кандидатом.
    :param user: объект user
    :return: bool
    """
    try:
        return user.member.is_slave()
    except AttributeError:
        return False


def check_booking_our_or_exception(pk, user):
    """Проверяет, что пользователь с айди анкеты pk был забронирован на направления пользователя user и
    рейзит ошибку, если не забронирован."""
    if not is_booked_by_user(pk, user):
        raise PermissionDenied('Данный пользователь не отобран на ваше направление.')


def is_booked_by_user(pk, user):
    """
    Возвращает True, если пользователь с айди анкеты = pk забронирован на направления user,
    в обратном случае - False
    """
    app = get_object_or_404(Application, pk=pk)
    try:
        master_affiliations = Affiliation.objects.filter(member=user.member)
    except AttributeError:
        return False
    return Booking.objects.filter(slave=app.member, booking_type__name=BOOKED,
                                  affiliation__in=master_affiliations).exists()


def add_additional_fields(request, user_app):
    additional_fields = [int(re.search('\d+', field).group(0)) for field in request.POST
                         if NAME_ADDITIONAL_FIELD_TEMPLATE in field]
    if additional_fields:
        addition_fields = AdditionField.objects.filter(pk__in=additional_fields)
        for field in addition_fields:
            AdditionFieldApp.objects.update_or_create(addition_field=field, application=user_app,
                                                      defaults={'value': request.POST.get(
                                                          f"{NAME_ADDITIONAL_FIELD_TEMPLATE}{field.id}")})


class WordTemplate:
    """ Класс для создания шаблона ворд документа по файлу, через путь path_to_template """

    def __init__(self, request, path_to_template):
        self.request = request
        self.path = path_to_template

    def create_word_in_buffer(self, context):
        """ Создает ворд документ и добавлет в него данные и сохраняет в буфер """
        template = DocxTemplate(self.path)
        user_docx = BytesIO()
        template.render(context=context)
        template.save(user_docx)
        user_docx.seek(0)
        return user_docx

    def create_context_to_interview_list(self, pk):
        """ Создает контекст для шаблона - 'Лист собеседования' """
        user_app = Application.objects.select_related('member').prefetch_related('education').defer('id').get(pk=pk)
        user_education = user_app.education.order_by('-end_year').values()
        user_education = user_education[0] if user_education else {}
        context = {**user_app.__dict__, **user_education}
        context.update({'father_name': user_app.member.father_name, 'phone': user_app.member.phone})
        context.update({'first_name': user_app.member.user.first_name, 'last_name': user_app.member.user.last_name})
        return context

    def create_context_to_word_files(self, document_type, all_directions=None):
        """ Создает контексты для шаблонов итоговых документов """
        current_year, current_season = get_current_draft_year()
        fixed_directions = self.request.user.member.affiliations.select_related('direction').all() \
            if not all_directions else Affiliation.objects.select_related('direction').all()

        context = {'directions': []}
        for direction in fixed_directions:
            platoon_data = {
                'name': direction.direction.name,
                'company_number': direction.company,
                'members': []
            }
            booked = Booking.objects.select_related('slave').filter(affiliation=direction, booking_type__name=BOOKED)
            booked_slaves = [b.slave for b in booked]
            booked_user_apps = Application.objects.select_related('scores', 'member__user').prefetch_related(
                'education'). \
                filter(member__in=booked_slaves, draft_year=current_year, draft_season=current_season[0]).all()

            for i, user_app in enumerate(booked_user_apps):
                try:
                    user_last_education = user_app.education.all()[0]
                except IndexError:
                    raise ValidationError(f'Файл не может быть сформирован, т.к. {user_app} не указал образование!')
                general_info, additional_info = {'number': i + 1,
                                                 'first_name': user_app.member.user.first_name,
                                                 'last_name': user_app.member.user.last_name,
                                                 'father_name': user_app.member.father_name,
                                                 'final_score': convert_float(user_app.final_score),
                                                 }, {}
                if document_type == PATH_TO_CANDIDATES_LIST:
                    additional_info = self._get_candidates_info(user_app, user_last_education)
                elif document_type == PATH_TO_RATING_LIST:
                    additional_info = self._get_rating_info(user_app, user_last_education)
                elif document_type == PATH_TO_EVALUATION_STATEMENT:
                    additional_info = self._get_evaluation_st_info(user_app)
                platoon_data['members'].append({**additional_info, **general_info})
            if platoon_data['members']:
                context['directions'].append(platoon_data)
        return context

    def create_context_to_psychological_test(self, user_test_result, questions, user_answers):
        """ Создает контекст для шаблона - 'Психологического теста ОПВС - 2' """
        user_app = Application.objects.only('birth_day').get(member=user_test_result.member)
        member = user_test_result.member
        context = {
            'full_name': f'{member.user.last_name} {member.user.first_name} {member.father_name}',
            'b_day': user_app.birth_day.strftime('%d %m %Y'),
            'now': datetime.datetime.now().strftime('%d %m %Y'),
            'position': 'гражданин',
            'questions': []
        }
        for i, question in enumerate(questions, 1):
            context['questions'].append({
                'num': i,
                'answers': [ans.id for ans in question.answer_options.all()],
                'response': user_answers.get(question.id)
            })
        return context

    def _get_evaluation_st_info(self, user_app):
        return {
            **{k: convert_float(v) for k, v in MEANING_COEFFICIENTS.items()},
            **{k: convert_float(v) for k, v in user_app.scores.__dict__.items() if isinstance(v, float)},
        }

    def _get_rating_info(self, user_app, user_last_education):
        return {
            'birth_day': user_app.birth_day,
            'military_commissariat': user_app.military_commissariat,
            'university': user_last_education.university,
            'specialization': user_last_education.specialization,
            'avg_score': convert_float(user_last_education.avg_score),
        }

    def _get_candidates_info(self, user_app, user_last_education):
        commissariat = MilitaryCommissariat.objects.filter(name=user_app.military_commissariat).first()
        return {
            'subject': commissariat.subject if commissariat else '',
            'birth_day': user_app.birth_day.year,
            'avg_score': convert_float(user_last_education.avg_score),
        }


def get_application_as_word(request, pk):
    """Генерирует word-файл анкеты сохраняет в буффер и возвращает"""
    word_template = WordTemplate(request, const.PATH_TO_INTERVIEW_LIST)
    context = word_template.create_context_to_interview_list(pk)
    return word_template.create_word_in_buffer(context)


def get_service_file(request, path_to_file, all_directions):
    """
    Генерирует сервисный файл, сохраняет в буффер и возвращает
    :param request: экземляр Request
    :param path_to_file: путь до шаблона
    :param all_directions: bool(нужно ли использовать все направления или только направления user'a)
    :return: file
    """
    word_template = WordTemplate(request, path_to_file)
    context = word_template.create_context_to_word_files(path_to_file, all_directions)
    return word_template.create_word_in_buffer(context)


def update_user_application_scores(pk):
    """Обновляет баллы анкеты с pk"""
    user_app = get_object_or_404(Application, pk=pk)
    user_app.update_scores(update_fields=['fullness', 'final_score'])


def set_is_final(application, value):
    """Разблокирует заявку для бронирования"""
    application.is_final = value
    application.save(update_fields=['is_final', ])


def set_work_group(application, value):
    """Устанавливает значение рабочей группы заявки"""
    application.work_group = value
    application.save(update_fields=["work_group"])


def get_booking(slave):
    """
    Возвращает экземпляр бронирования заявки
    :param slave: экземпляр Member
    :return: экземпляр Booking переданного slave или False
    """
    booking = Booking.objects.filter(slave=slave, booking_type=get_booked_type())
    return booking.first() if booking else False


def get_competence_list(direction_id, picked):
    """
    Возвращает список компетенций направления если picked=True, иначе возвращает список невыбранных компетенций
    на данное направление
    :param direction_id: идентификатор направления
    :param picked: Bool - нужно выбирать только выбранные компетенции или все, кроме выбранных
    :return: список компетенций(queryset)
    """
    if picked:
        return Competence.objects.filter(directions__id=direction_id)
    return Competence.objects.exclude(directions__id=direction_id)


def parse_str_to_bool(string):
    """Конвертирует входную строку в bool"""
    try:
        if string in TRUE_VALUES:
            return True
        elif string in FALSE_VALUES:
            return False
    except TypeError:
        raise ParseError('Плохой query параметр')


def remove_direction_from_competence_list(direction_id, competences_id):
    """
    Удаляет направление из списка компетенции
    :param direction_id: id направления
    :param competences_id: список id компетенций
    :return:
    """
    with transaction.atomic():
        for competence in Competence.objects.filter(pk__in=competences_id):
            competence.directions.remove(direction_id)


def add_direction_to_competence_list(direction_id, competences_id):
    """
    Добавляет направление в список компетенции
    :param direction_id: id направления
    :param competences_id: список id компетенций
    :return:
    """
    with transaction.atomic():
        for competence in Competence.objects.filter(pk__in=competences_id):
            competence.directions.add(direction_id)


def has_application_viewed(application, member):
    """
    Проверяет, была ли заявка уже просмотрена
    :param application: объект Application
    :param member: объект Member
    :return: True/False
    """
    return ViewedApplication.objects.filter(member=member, application=application).exists()


class NumberInFilter(BaseInFilter, NumberFilter):
    pass


class ApplicationFilter(FilterSet):
    """Фильтр анкет."""
    draft_year = AllValuesMultipleFilter(field_name='draft_year')
    directions = NumberInFilter(field_name='directions__id', lookup_expr='in')
    draft_season = NumberInFilter(field_name='draft_season', lookup_expr='in')
    booking_aff = NumberInFilter(label='Отобрано во взвод',
                                 method='filter_booking_aff')  # id affiliation, на которые отобраны заявки
    wishlist_aff = NumberInFilter(label='Избранное во взвод',
                                  method='filter_wishlist_aff')  # id affiliation, для которых заявки добавлены в вишлист

    def filter_booking_aff(self, queryset, name, value):
        """
        Фильтрует queryset.

        Оставялет только те заявки, которые были отобраны на принадлежности, переданные в value.
        :param queryset: исходный queryset
        :param name: имя query-параметра
        :param value: список affiliations id
        :return: отфильтрованный queryset
        """
        booked_members = Booking.objects.filter(affiliation__in=value, booking_type__name=const.BOOKED) \
            .values_list('slave', flat=True)
        return queryset.filter(member_id__in=booked_members).distinct()

    def filter_wishlist_aff(self, queryset, name, value):
        """
        Фильтрует queryset.

        Оставялет только те заявки, которые были добавленны в избранное на принадлежности, переданные в value.
        :param queryset: исходный queryset
        :param name: имя query-параметра
        :param value: список affiliations id
        :return: отфильтрованный queryset
        """
        wish_list_members = Booking.objects.filter(affiliation__in=value, booking_type__name=const.IN_WISHLIST) \
            .values_list('slave', flat=True)
        return queryset.filter(member__id__in=wish_list_members).distinct()

    class Meta:
        model = Application
        fields = ('directions', 'booking_aff', 'wishlist_aff', 'draft_season', 'draft_year')


class CustomOrderingFilter(OrderingFilter):
    """Фильтр со стандартной сортировкой анкет"""

    def filter_queryset(self, request, queryset, view):
        # Сортирует queryset.
        # Default ordering применяется всегда. Остальные сортировки происходят после default.
        # Сортирует только если пользователь является отбирающим.
        if not is_master(request.user):
            return queryset
        ordering = super().get_default_ordering(view)
        ordering.extend(self.get_ordering(request, queryset, view))

        if ordering:
            return queryset.order_by(*ordering)

        return queryset


class ApplicationExporter:
    """Экспорт списка заявок в exel"""

    def __init__(self, applications):
        """
        Устанавливает список анкет, создает и устанавливает книгу и лист.
        :param applications: queryset заявок
        """
        self.applications = applications
        self.wb = Workbook(write_only=True)
        self.sheet = self.wb.create_sheet()

    def add_applications_to_sheet(self):
        """Добавляет заявки в файл и сохраняет его."""
        header = const.HEADERS_FOR_EXCEL_APP_TABLES
        self._set_column_dimensions(header)
        self.sheet.append(header)
        for app in self.applications:
            row = self._convert_applications_to_required_format(app)
            self.sheet.append(row)
        return self._save()

    def add_work_list_to_sheet(self):
        """Добавляет заявки забочего листа в файл и сохраняет."""
        header = const.WORK_LIST_HEADERS_FOR_EXCEL
        self._set_column_dimensions(header)
        self.sheet.append(header)
        for app in self.applications:
            row = self._convert_work_list_to_required_format(app)
            self.sheet.append(row)
        return self._save()

    def _convert_applications_to_required_format(self, app):
        """Конвертирует заявки в нужный формат."""
        birth_day = datetime.datetime.strftime(app.birth_day, '%d.%m.%Y')
        draft_season = app.get_draft_time()
        full_name = self._get_full_name(app)
        university, education_type, specialization, avg_score = self._get_education_info(app)
        return [full_name, draft_season, birth_day, app.birth_place, app.subject,
                university, education_type, specialization, avg_score]

    def _get_education_info(self, application):
        """Возвращает информацию об образовании кандидата, если оно существует."""
        education = application.education.first()
        return (education.university, education.get_education_type_display(), education.specialization,
                education.avg_score) if education else ('', '', '', '')

    def _convert_work_list_to_required_format(self, app):
        """Конвертирует заявки рабочего листа в нужный формат."""
        full_name = self._get_full_name(app)
        user_competencies = self._get_user_competencies(app)
        return [full_name, app.member.phone, app.member.user.email, app.final_score, app.university, app.specialization,
                ', '.join(user_competencies[3]), ', '.join(user_competencies[2]), ', '.join(user_competencies[1])]

    def _get_user_competencies(self, app):
        """Преобразовывает компетенции заявок в нужный формат и возвращает их."""
        competence_levels = {
            3: [],
            2: [],
            1: [],
        }
        for comp in app.app_competence.all():
            competence_levels[comp.level].append(comp.competence.name)
        return competence_levels

    def _get_full_name(self, app):
        """Возвращает полное имя кандидата в заявке."""
        return f"{app.member.user.last_name} {app.member.user.first_name} {app.member.father_name}"

    def _save(self):
        """Сохраняет файл."""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _set_column_dimensions(self, columns):
        """Устанавливает ширину колонок."""
        for i in range(1, len(columns) + 1):
            letter = get_column_letter(i)
            self.sheet.column_dimensions[letter].width = 30


def get_applications_by_master(user, master_affiliations, master_directions, master_directions_id):
    """
    Возвращает queryset заявок с аннотированными полями.

    Переданный user должен иметь роль master.
    :param user: экземляр user(мастер)
    :param master_affiliations: список принадлежностей мастера
    :param master_directions: список направлений мастера
    :param master_directions_id: список id направлений мастера
    :return: queryset(Application)
    """
    apps = (
        Application.objects.all()
            .select_related("member", "member__user")
            .prefetch_related(
            "education",
            "directions",
            Prefetch(
                "notes",
                queryset=ApplicationNote.objects.filter(
                    author=user.member,
                    affiliations__in=master_affiliations,
                )
                    .select_related("author__user")
                    .prefetch_related("affiliations"),
            ),
            Prefetch(
                "member__candidate",
                queryset=Booking.objects.filter(
                    affiliation__in=master_affiliations,
                    booking_type__name=const.IN_WISHLIST,
                ).select_related("affiliation", "master__user"),
            ),
            Prefetch(
                "member__candidate",
                queryset=Booking.objects.filter(
                    booking_type__name=const.BOOKED
                ).select_related("affiliation", "master__user"),
                to_attr="booking_affiliation",
            ),
            Prefetch(
                "directions",
                queryset=master_directions,
                to_attr="available_booking_direction",
            ),
        )
            .annotate(
            is_booked=Count(
                F("member__candidate"),
                filter=Q(member__candidate__booking_type__name=const.BOOKED),
                distinct=True,
            ),
            is_booked_our=Count(
                F("member__candidate"),
                filter=Q(
                    member__candidate__booking_type__name=const.BOOKED,
                    member__candidate__affiliation__in=master_affiliations,
                ),
                distinct=True,
            ),
            can_unbook=Count(
                F("member__candidate"),
                filter=Q(
                    member__candidate__booking_type__name=const.BOOKED,
                    member__candidate__affiliation__in=master_affiliations,
                    member__candidate__master=user.member,
                ),
                distinct=True,
            ),
            wishlist_len=Count(
                F("member__candidate"),
                filter=Q(member__candidate__booking_type__name=const.IN_WISHLIST),
                distinct=True,
            ),
            is_in_wishlist=Count(
                F("member__candidate"),
                filter=Q(
                    member__candidate__booking_type__name=const.IN_WISHLIST,
                    member__candidate__affiliation__in=master_affiliations,
                ),
                distinct=True,
            ),
            our_direction_count=Count(
                F('directions'),
                filter=Q(directions__id__in=master_directions_id),
                distinct=True
            ),
            our_direction=Case(
                When(our_direction_count__gt=0, then=Value(True)),
                default=Value(False),
            ),
            subject=(
                MilitaryCommissariat.objects.filter(
                    name=OuterRef("military_commissariat")
                ).values_list("subject")[:1]
            ),
            is_viewed=Count(
                F("viewed"),
                filter=Q(viewed__member=user.member),
                distinct=True,
            ),
        )
    )
    return apps


def get_applications_by_slave():
    """
    Возвращает queryset заявок с аннотированными полями.

    :return: queryset(Application)
    """
    apps = (
        Application.objects.all()
            .select_related("member", "member__user")
            .prefetch_related(
            "education",
            "directions",
            Prefetch(
                "member__candidate",
                queryset=Booking.objects.filter(
                    booking_type__name=const.BOOKED
                ).select_related("affiliation", "master__user"),
                to_attr="booking_affiliation",
            ),
        )
            .annotate(
            is_booked=Count(
                F("member__candidate"),
                filter=Q(member__candidate__booking_type__name=const.BOOKED),
                distinct=True,
            ),
            subject=(
                MilitaryCommissariat.objects.filter(
                    name=OuterRef("military_commissariat")
                ).values_list("subject")[:1]
            ),
        )
    )
    return apps
